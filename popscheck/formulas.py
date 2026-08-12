"""Static, opt-in analysis of Excel formulas.

This module intentionally has no dependency on :mod:`popscheck.compare`,
:mod:`popscheck.models`, or :mod:`popscheck.reporting`.  It can therefore be
used by callers that want a formula audit without changing the existing
structural-comparison API.

The parser is deliberately conservative.  ``openpyxl``'s formula tokenizer is
used to avoid interpreting references inside string literals.  Ordinary A1
references and row/column ranges are understood; defined names, structured
references, external references, and 3-D references are retained as opaque
tokens rather than guessed at.
"""

from __future__ import annotations

from collections import defaultdict
from contextlib import contextmanager
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
import re
from types import MappingProxyType
from typing import Any, Iterable, Iterator, Mapping

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula import Tokenizer
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string


EXCEL_MAX_ROW = 1_048_576
EXCEL_MAX_COLUMN = 16_384

_CELL_RE = re.compile(
    r"^(?P<column_absolute>\$?)(?P<column>[A-Z]{1,3})"
    r"(?P<row_absolute>\$?)(?P<row>[1-9]\d*)$",
    re.IGNORECASE,
)
_COLUMN_RE = re.compile(
    r"^(?P<absolute>\$?)(?P<column>[A-Z]{1,3})$", re.IGNORECASE
)
_ROW_RE = re.compile(r"^(?P<absolute>\$?)(?P<row>[1-9]\d*)$")
_EXTERNAL_QUALIFIER_RE = re.compile(
    r"^(?P<path>.*)\[(?P<book>[^\]]+)\](?P<sheet>.*)$"
)


class MappingDirection(str, Enum):
    """Direction in which physical coordinates are rewritten."""

    NONE = "none"
    EXPECTED_TO_OBSERVED = "expected_to_observed"
    OBSERVED_TO_EXPECTED = "observed_to_expected"


class FormulaChangeKind(str, Enum):
    """Kinds emitted by :func:`compare_formula_inventories`."""

    ADDED = "formula_added"
    REMOVED = "formula_removed"
    REPLACED = "formula_replaced"
    MODIFIED = "formula_modified"
    BROKEN_REFERENCE = "broken_reference"


@dataclass(frozen=True, order=True, slots=True)
class CellAddress:
    """A cell address with an explicit worksheet name."""

    sheet: str
    row: int
    column: int

    def __post_init__(self) -> None:
        if not self.sheet:
            raise ValueError("sheet must not be empty")
        if not 1 <= self.row <= EXCEL_MAX_ROW:
            raise ValueError(f"row outside Excel limits: {self.row}")
        if not 1 <= self.column <= EXCEL_MAX_COLUMN:
            raise ValueError(f"column outside Excel limits: {self.column}")

    @property
    def coordinate(self) -> str:
        return f"{get_column_letter(self.column)}{self.row}"

    @property
    def qualified(self) -> str:
        return f"{_quote_sheet(self.sheet)}!{self.coordinate}"


@dataclass(frozen=True, slots=True)
class SheetCoordinateMapping:
    """Expected-to-observed row and column mappings for one worksheet.

    ``None`` means that an axis is unchanged and should use identity mapping.
    A provided mapping is authoritative: an expected coordinate absent from it
    is considered removed, and an observed coordinate absent from its inverse
    is considered added.  Mappings must be one-to-one.
    """

    rows: Mapping[int, int] | None = None
    columns: Mapping[int, int] | None = None
    _inverse_rows: Mapping[int, int] | None = field(
        default=None, init=False, repr=False, compare=False
    )
    _inverse_columns: Mapping[int, int] | None = field(
        default=None, init=False, repr=False, compare=False
    )

    def __post_init__(self) -> None:
        for name, mapping, limit in (
            ("rows", self.rows, EXCEL_MAX_ROW),
            ("columns", self.columns, EXCEL_MAX_COLUMN),
        ):
            if mapping is None:
                continue
            copied = dict(mapping)
            if any(
                isinstance(key, bool)
                or isinstance(value, bool)
                or not isinstance(key, int)
                or not isinstance(value, int)
                or not 1 <= key <= limit
                or not 1 <= value <= limit
                for key, value in copied.items()
            ):
                raise ValueError(f"{name} mapping contains an invalid coordinate")
            if len(set(copied.values())) != len(copied):
                raise ValueError(f"{name} mapping must be one-to-one")
            frozen = MappingProxyType(copied)
            inverse = MappingProxyType(_inverse(copied))
            object.__setattr__(self, name, frozen)
            object.__setattr__(self, f"_inverse_{name}", inverse)

    def expected_row(self, row: int) -> int | None:
        return row if self.rows is None else self.rows.get(row)

    def expected_column(self, column: int) -> int | None:
        return column if self.columns is None else self.columns.get(column)

    def observed_row(self, row: int) -> int | None:
        if self.rows is None:
            return row
        assert self._inverse_rows is not None
        return self._inverse_rows.get(row)

    def observed_column(self, column: int) -> int | None:
        if self.columns is None:
            return column
        assert self._inverse_columns is not None
        return self._inverse_columns.get(column)


@dataclass(frozen=True, slots=True)
class FormulaReference:
    """A reference token found in a formula.

    Coordinates are populated for understood A1 references.  ``kind`` is one
    of ``cell``, ``cell_range``, ``row_range``, ``column_range``, ``opaque``,
    ``three_dimensional``, or ``broken``.
    """

    raw: str
    normalized: str
    kind: str
    sheet: str | None = None
    explicit_sheet: bool = False
    external_book: str | None = None
    start_row: int | None = None
    start_column: int | None = None
    end_row: int | None = None
    end_column: int | None = None
    unresolved_mapping: bool = False

    @property
    def is_external(self) -> bool:
        return self.external_book is not None

    @property
    def is_broken(self) -> bool:
        return self.kind == "broken" or "#REF!" in self.raw.upper()


@dataclass(frozen=True, slots=True)
class ParsedFormula:
    """Tokenized and conservatively normalized formula."""

    raw: str
    normalized: str
    references: tuple[FormulaReference, ...] = ()
    has_ref_error: bool = False
    warnings: tuple[str, ...] = ()
    address: CellAddress | None = None


@dataclass(slots=True)
class FormulaInventory:
    """Formula and literal-cell inventory for a workbook."""

    formulas: dict[CellAddress, ParsedFormula] = field(default_factory=dict)
    literal_cells: frozenset[CellAddress] = frozenset()
    sheet_names: tuple[str, ...] = ()
    defined_names: frozenset[str] = frozenset()
    table_names: frozenset[str] = frozenset()
    defined_name_definitions: dict[str, str] = field(default_factory=dict)
    table_definitions: dict[str, str] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    def dependencies(self) -> tuple["SheetDependency", ...]:
        return build_inter_sheet_dependencies(self)


@dataclass(frozen=True, slots=True)
class FormulaChange:
    """One formula-level difference or broken-reference finding."""

    kind: FormulaChangeKind
    expected_address: CellAddress | None = None
    observed_address: CellAddress | None = None
    expected_formula: str | None = None
    observed_formula: str | None = None
    expected_normalized: str | None = None
    observed_normalized: str | None = None
    message: str = ""

    @property
    def address(self) -> CellAddress | None:
        return self.observed_address or self.expected_address


@dataclass(frozen=True, order=True, slots=True)
class SheetDependency:
    """Aggregated static dependency from one worksheet to another."""

    source_sheet: str
    target_sheet: str
    formula_cells: tuple[CellAddress, ...] = ()
    reference_count: int = 0
    target_exists: bool = True


@dataclass(frozen=True, slots=True)
class _CellPart:
    row: int
    column: int
    row_absolute: bool
    column_absolute: bool


def _inverse(mapping: Mapping[int, int]) -> dict[int, int]:
    return {value: key for key, value in mapping.items()}


def _coerce_direction(value: MappingDirection | str) -> MappingDirection:
    if isinstance(value, MappingDirection):
        return value
    try:
        return MappingDirection(value)
    except ValueError as exc:
        choices = ", ".join(item.value for item in MappingDirection)
        raise ValueError(f"mapping_direction must be one of: {choices}") from exc


def _formula_text(value: Any) -> str:
    if hasattr(value, "text"):
        value = value.text
    text = str(value or "")
    if text and not text.startswith("="):
        text = f"={text}"
    return text


def _quote_sheet(sheet: str) -> str:
    return "'" + sheet.replace("'", "''") + "'"


def _unquote_sheet(value: str) -> str:
    stripped = value.strip()
    if len(stripped) >= 2 and stripped[0] == stripped[-1] == "'":
        return stripped[1:-1].replace("''", "'")
    return stripped


def _known_sheet_name(sheet: str, known_sheets: Iterable[str] | None) -> str:
    if known_sheets is None:
        return sheet
    by_folded = {candidate.casefold(): candidate for candidate in known_sheets}
    return by_folded.get(sheet.casefold(), sheet)


def _mapping_for(
    sheet: str | None,
    mappings: Mapping[str, SheetCoordinateMapping] | None,
) -> SheetCoordinateMapping | None:
    if sheet is None or not mappings:
        return None
    direct = mappings.get(sheet)
    if direct is not None:
        return direct
    folded = sheet.casefold()
    return next(
        (mapping for name, mapping in mappings.items() if name.casefold() == folded),
        None,
    )


def _map_axis(
    value: int,
    mapping: SheetCoordinateMapping | None,
    axis: str,
    direction: MappingDirection,
) -> tuple[int, bool]:
    if mapping is None or direction is MappingDirection.NONE:
        return value, False
    if axis == "row":
        translated = (
            mapping.expected_row(value)
            if direction is MappingDirection.EXPECTED_TO_OBSERVED
            else mapping.observed_row(value)
        )
    elif axis == "column":
        translated = (
            mapping.expected_column(value)
            if direction is MappingDirection.EXPECTED_TO_OBSERVED
            else mapping.observed_column(value)
        )
    else:  # defensive guard for internal callers
        raise ValueError(f"unknown axis: {axis}")
    if translated is None:
        translated = _project_outside_mapping(value, mapping, axis, direction)
    if translated is None:
        # Preserving the coordinate is safer than inventing one.  Callers can
        # inspect ``unresolved_mapping`` and decide whether to escalate it.
        return value, True
    return translated, False


def _project_outside_mapping(
    value: int,
    mapping: SheetCoordinateMapping,
    axis: str,
    direction: MappingDirection,
) -> int | None:
    """Project references beyond the finite structural scan envelope.

    Axis mappings are deliberately authoritative for formula *hosts*: a
    missing coordinate means an inserted or deleted row/column.  Formula
    references, however, may legitimately point far beyond the populated
    extent (for example ``A100``).  Excel shifts those references after an
    insertion near the top.  Outside the mapped envelope only, continue the
    offset of the nearest boundary anchor; gaps inside the envelope remain
    unresolved and are never guessed.
    """

    configured = mapping.rows if axis == "row" else mapping.columns
    if not configured:
        return None
    if direction is MappingDirection.EXPECTED_TO_OBSERVED:
        source_to_target = configured
    else:
        source_to_target = {target: source for source, target in configured.items()}
    lower = min(source_to_target)
    upper = max(source_to_target)
    if lower <= value <= upper:
        return None
    anchor = lower if value < lower else upper
    projected = source_to_target[anchor] + (value - anchor)
    limit = EXCEL_MAX_ROW if axis == "row" else EXCEL_MAX_COLUMN
    return projected if 1 <= projected <= limit else None


def _parse_cell(value: str) -> _CellPart | None:
    match = _CELL_RE.fullmatch(value)
    if not match:
        return None
    row = int(match.group("row"))
    try:
        column = column_index_from_string(match.group("column").upper())
    except ValueError:
        return None
    if row > EXCEL_MAX_ROW or column > EXCEL_MAX_COLUMN:
        return None
    return _CellPart(
        row=row,
        column=column,
        row_absolute=bool(match.group("row_absolute")),
        column_absolute=bool(match.group("column_absolute")),
    )


def _render_cell(part: _CellPart, row: int, column: int) -> str:
    return (
        ("$" if part.column_absolute else "")
        + get_column_letter(column)
        + ("$" if part.row_absolute else "")
        + str(row)
    )


def _split_qualifier(value: str) -> tuple[str | None, str | None, str, bool]:
    """Return ``sheet, external_book, reference_body, explicit_sheet``."""

    if "!" not in value:
        return None, None, value, False
    qualifier, body = value.rsplit("!", 1)
    qualifier = _unquote_sheet(qualifier)
    external_book: str | None = None
    external = _EXTERNAL_QUALIFIER_RE.match(qualifier)
    if external:
        external_book = external.group("path") + external.group("book")
        qualifier = external.group("sheet")
    return qualifier, external_book, body, True


def _canonical_qualifier(
    sheet: str | None,
    external_book: str | None,
    explicit: bool,
) -> str:
    if not explicit or sheet is None:
        return ""
    qualified = f"[{external_book}]{sheet}" if external_book else sheet
    return f"{_quote_sheet(qualified)}!"


def _reference_from_token(
    token_value: str,
    *,
    origin: CellAddress | None,
    mappings: Mapping[str, SheetCoordinateMapping] | None,
    direction: MappingDirection,
    known_sheets: Iterable[str] | None,
) -> FormulaReference | None:
    raw = token_value
    upper_raw = raw.upper()
    if "#REF!" in upper_raw:
        return FormulaReference(
            raw=raw,
            normalized=upper_raw,
            kind="broken",
            explicit_sheet="!" in raw,
        )

    sheet, external_book, body, explicit = _split_qualifier(raw)
    if sheet is not None:
        sheet = _known_sheet_name(sheet, known_sheets)
    effective_sheet = sheet or (origin.sheet if origin else None)
    qualifier = _canonical_qualifier(sheet, external_book, explicit)

    if sheet is not None and ":" in sheet:
        return FormulaReference(
            raw=raw,
            normalized=qualifier + body.upper(),
            kind="three_dimensional",
            sheet=sheet,
            explicit_sheet=explicit,
            external_book=external_book,
        )

    prefix = "@" if body.startswith("@") else ""
    suffix = "#" if body.endswith("#") else ""
    core = body[len(prefix) : len(body) - len(suffix) if suffix else None]
    parts = core.split(":")
    mapping = _mapping_for(effective_sheet, mappings)

    if len(parts) in {1, 2}:
        cells = [_parse_cell(part) for part in parts]
        if all(cell is not None for cell in cells):
            rendered: list[str] = []
            unresolved = False
            for cell in cells:
                assert cell is not None
                row, row_unresolved = _map_axis(
                    cell.row, mapping, "row", direction
                )
                column, column_unresolved = _map_axis(
                    cell.column, mapping, "column", direction
                )
                unresolved |= row_unresolved or column_unresolved
                rendered.append(_render_cell(cell, row, column))
            normalized_body = ":".join(rendered)
            first = cells[0]
            last = cells[-1]
            assert first is not None and last is not None
            return FormulaReference(
                raw=raw,
                normalized=qualifier + prefix + normalized_body + suffix,
                kind="cell" if len(cells) == 1 else "cell_range",
                sheet=effective_sheet,
                explicit_sheet=explicit,
                external_book=external_book,
                start_row=first.row,
                start_column=first.column,
                end_row=last.row,
                end_column=last.column,
                unresolved_mapping=unresolved,
            )

    if len(parts) == 2:
        column_parts = [_COLUMN_RE.fullmatch(part) for part in parts]
        if all(match is not None for match in column_parts):
            rendered_columns: list[str] = []
            unresolved = False
            original_columns: list[int] = []
            for match in column_parts:
                assert match is not None
                try:
                    column = column_index_from_string(match.group("column").upper())
                except ValueError:
                    break
                if column > EXCEL_MAX_COLUMN:
                    break
                original_columns.append(column)
                mapped, missed = _map_axis(
                    column, mapping, "column", direction
                )
                unresolved |= missed
                rendered_columns.append(
                    ("$" if match.group("absolute") else "")
                    + get_column_letter(mapped)
                )
            if len(rendered_columns) == 2:
                return FormulaReference(
                    raw=raw,
                    normalized=qualifier + ":".join(rendered_columns),
                    kind="column_range",
                    sheet=effective_sheet,
                    explicit_sheet=explicit,
                    external_book=external_book,
                    start_column=original_columns[0],
                    end_column=original_columns[1],
                    unresolved_mapping=unresolved,
                )

        row_parts = [_ROW_RE.fullmatch(part) for part in parts]
        if all(match is not None for match in row_parts):
            rendered_rows: list[str] = []
            unresolved = False
            original_rows: list[int] = []
            for match in row_parts:
                assert match is not None
                row = int(match.group("row"))
                if row > EXCEL_MAX_ROW:
                    break
                original_rows.append(row)
                mapped, missed = _map_axis(
                    row, mapping, "row", direction
                )
                unresolved |= missed
                rendered_rows.append(
                    ("$" if match.group("absolute") else "") + str(mapped)
                )
            if len(rendered_rows) == 2:
                return FormulaReference(
                    raw=raw,
                    normalized=qualifier + ":".join(rendered_rows),
                    kind="row_range",
                    sheet=effective_sheet,
                    explicit_sheet=explicit,
                    external_book=external_book,
                    start_row=original_rows[0],
                    end_row=original_rows[1],
                    unresolved_mapping=unresolved,
                )

    # Defined names and structured references are valid RANGE operands too.  We
    # preserve them and still retain an explicit sheet for dependency analysis.
    return FormulaReference(
        raw=raw,
        normalized=qualifier + body.upper(),
        kind="opaque",
        sheet=effective_sheet,
        explicit_sheet=explicit,
        external_book=external_book,
    )


def parse_formula(
    formula: Any,
    *,
    address: CellAddress | None = None,
    mappings: Mapping[str, SheetCoordinateMapping] | None = None,
    mapping_direction: MappingDirection | str = MappingDirection.NONE,
    known_sheets: Iterable[str] | None = None,
) -> ParsedFormula:
    """Tokenize and normalize one Excel formula without evaluating it.

    Invalid or unsupported syntax is returned as a conservative raw fallback
    with a warning instead of escaping as a tokenizer exception.
    """

    raw = _formula_text(formula)
    if not raw:
        raise ValueError("formula must not be empty")
    direction = _coerce_direction(mapping_direction)
    references: list[FormulaReference] = []
    warnings: list[str] = []
    normalized_parts: list[str] = []
    has_ref_error = False
    try:
        tokenizer = Tokenizer(raw)
        tokens = list(tokenizer.items)
        for index, token in enumerate(tokens):
            value = token.value
            if token.type == "OPERAND" and token.subtype == "RANGE":
                reference = _reference_from_token(
                    value,
                    origin=address,
                    mappings=mappings,
                    direction=direction,
                    known_sheets=known_sheets,
                )
                if reference is not None:
                    references.append(reference)
                    value = reference.normalized
                    has_ref_error |= reference.is_broken
            elif token.type == "OPERAND" and token.subtype == "ERROR":
                value = value.upper()
                has_ref_error |= value == "#REF!"
            elif token.type == "WHITE-SPACE":
                # Whitespace is semantically meaningful only as Excel's range
                # intersection operator. Formatting around +, -, commas, etc.
                # must not create a formula-change false positive.
                previous = next(
                    (
                        item
                        for item in reversed(tokens[:index])
                        if item.type != "WHITE-SPACE"
                    ),
                    None,
                )
                following = next(
                    (
                        item
                        for item in tokens[index + 1 :]
                        if item.type != "WHITE-SPACE"
                    ),
                    None,
                )
                if (
                    previous is not None
                    and following is not None
                    and previous.type == following.type == "OPERAND"
                    and previous.subtype == following.subtype == "RANGE"
                ):
                    value = " " if not normalized_parts or normalized_parts[-1] != " " else ""
                else:
                    value = ""
            elif not (token.type == "OPERAND" and token.subtype == "TEXT"):
                # Function and name spelling is case-insensitive in Excel;
                # string literal contents are not.
                value = value.upper()
            normalized_parts.append(value)
    except Exception as exc:  # pragma: no cover - tokenizer internals vary
        warnings.append(f"Formula tokenizer fallback: {type(exc).__name__}: {exc}")
        normalized_parts = [raw[1:] if raw.startswith("=") else raw]
        # With no token stream we cannot distinguish an error token from a text
        # literal.  Report this uncertainty instead of asserting a #REF! issue.
        has_ref_error = False

    for reference in references:
        if reference.unresolved_mapping:
            warnings.append(
                f"Unmapped coordinate retained in reference {reference.raw!r}."
            )
        if reference.kind == "three_dimensional":
            warnings.append(
                f"3-D reference retained without expansion: {reference.raw!r}."
            )
    return ParsedFormula(
        raw=raw,
        normalized="=" + "".join(normalized_parts),
        references=tuple(references),
        has_ref_error=has_ref_error,
        warnings=tuple(dict.fromkeys(warnings)),
        address=address,
    )


def normalize_formula_references(
    formula: Any,
    *,
    address: CellAddress | None = None,
    mappings: Mapping[str, SheetCoordinateMapping] | None = None,
    mapping_direction: MappingDirection | str = MappingDirection.NONE,
    known_sheets: Iterable[str] | None = None,
) -> str:
    """Return only the normalized text produced by :func:`parse_formula`."""

    return parse_formula(
        formula,
        address=address,
        mappings=mappings,
        mapping_direction=mapping_direction,
        known_sheets=known_sheets,
    ).normalized


def _iter_cells(worksheet: Any) -> Iterator[Any]:
    cells = getattr(worksheet, "_cells", None)
    if isinstance(cells, dict):
        for cell in cells.values():
            if not isinstance(cell, MergedCell):
                yield cell
        return
    for row in worksheet.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                yield cell


@contextmanager
def _workbook_source(source: Any) -> Iterator[Any]:
    if isinstance(source, (str, Path)):
        path = Path(source)
        keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
        workbook = load_workbook(
            path,
            read_only=False,
            data_only=False,
            keep_vba=keep_vba,
            keep_links=True,
            rich_text=False,
        )
        try:
            yield workbook
        finally:
            workbook.close()
        return
    if not hasattr(source, "worksheets"):
        raise TypeError("source must be a workbook or an Excel workbook path")
    yield source


def extract_formulas(
    source: Any,
    *,
    mappings: Mapping[str, SheetCoordinateMapping] | None = None,
    mapping_direction: MappingDirection | str = MappingDirection.NONE,
) -> FormulaInventory:
    """Extract formulas and materialized literal cells from a workbook or path."""

    direction = _coerce_direction(mapping_direction)
    with _workbook_source(source) as workbook:
        sheet_names = tuple(workbook.sheetnames)
        defined_names: set[str] = set()
        defined_name_definitions: dict[str, str] = {}
        try:
            for item in workbook.defined_names.values():
                if not getattr(item, "name", None):
                    continue
                name = str(item.name)
                defined_names.add(name)
                defined_name_definitions[name] = str(
                    getattr(item, "attr_text", "") or ""
                )
        except (AttributeError, TypeError):
            pass
        table_names: set[str] = set()
        table_definitions: dict[str, str] = {}
        formulas: dict[CellAddress, ParsedFormula] = {}
        literal_cells: set[CellAddress] = set()
        warnings: list[str] = []
        for worksheet in workbook.worksheets:
            try:
                for name in worksheet.tables.keys():
                    table = worksheet.tables[name]
                    rendered_name = str(name)
                    table_names.add(rendered_name)
                    table_definitions[rendered_name] = (
                        f"'{worksheet.title}'!{getattr(table, 'ref', '')}"
                    )
            except (AttributeError, TypeError):
                pass
            for cell in _iter_cells(worksheet):
                value = cell.value
                if value is None:
                    continue
                address = CellAddress(worksheet.title, cell.row, cell.column)
                # ``openpyxl`` marks normal, shared, and array formulas with
                # data_type ``f``.  Trust that metadata: a literal text cell is
                # allowed to begin with ``=`` and must not be promoted merely
                # because of its spelling.
                is_formula = cell.data_type == "f"
                if not is_formula:
                    literal_cells.add(address)
                    continue
                parsed = parse_formula(
                    value,
                    address=address,
                    mappings=mappings,
                    mapping_direction=direction,
                    known_sheets=sheet_names,
                )
                formulas[address] = parsed
                warnings.extend(
                    f"{address.qualified}: {warning}" for warning in parsed.warnings
                )
        return FormulaInventory(
            formulas=formulas,
            literal_cells=frozenset(literal_cells),
            sheet_names=sheet_names,
            defined_names=frozenset(defined_names),
            table_names=frozenset(table_names),
            defined_name_definitions=defined_name_definitions,
            table_definitions=table_definitions,
            warnings=list(dict.fromkeys(warnings)),
        )


def _address_lookup(addresses: Iterable[CellAddress]) -> dict[tuple[str, int, int], CellAddress]:
    return {
        (address.sheet.casefold(), address.row, address.column): address
        for address in addresses
    }


def _mapped_observed_address(
    address: CellAddress,
    mappings: Mapping[str, SheetCoordinateMapping] | None,
) -> CellAddress | None:
    mapping = _mapping_for(address.sheet, mappings)
    if mapping is None:
        return address
    row = mapping.expected_row(address.row)
    column = mapping.expected_column(address.column)
    if row is None or column is None:
        return None
    return CellAddress(address.sheet, row, column)


def _change_message(
    kind: FormulaChangeKind,
    expected: CellAddress | None,
    observed: CellAddress | None,
) -> str:
    location = (observed or expected).qualified if (observed or expected) else "unknown cell"
    messages = {
        FormulaChangeKind.ADDED: f"Formula added at {location}.",
        FormulaChangeKind.REMOVED: f"Formula removed at {location}.",
        FormulaChangeKind.REPLACED: f"Formula replaced by a literal value at {location}.",
        FormulaChangeKind.MODIFIED: f"Formula modified at {location}.",
        FormulaChangeKind.BROKEN_REFERENCE: f"Formula contains #REF! at {location}.",
    }
    return messages[kind]


def compare_formula_inventories(
    expected: FormulaInventory,
    observed: FormulaInventory,
    *,
    mappings: Mapping[str, SheetCoordinateMapping] | None = None,
) -> tuple[FormulaChange, ...]:
    """Compare two inventories in expected/canonical coordinates.

    Formula replacement means that a formula cell now contains a non-formula
    literal.  A blank cell is classified as removal.  A formula containing
    ``#REF!`` produces a separate ``BROKEN_REFERENCE`` finding, so callers do
    not lose the fact that the formula was also modified or added.
    """

    observed_formula_lookup = _address_lookup(observed.formulas)
    observed_literal_lookup = _address_lookup(observed.literal_cells)
    matched_observed: set[CellAddress] = set()
    changes: list[FormulaChange] = []
    known_sheets = tuple(dict.fromkeys(expected.sheet_names + observed.sheet_names))

    for expected_address, expected_formula in sorted(expected.formulas.items()):
        requested_observed = _mapped_observed_address(expected_address, mappings)
        observed_address: CellAddress | None = None
        if requested_observed is not None:
            key = (
                requested_observed.sheet.casefold(),
                requested_observed.row,
                requested_observed.column,
            )
            observed_address = observed_formula_lookup.get(key)
            literal_address = observed_literal_lookup.get(key)
        else:
            literal_address = None

        expected_parsed = parse_formula(
            expected_formula.raw,
            address=expected_address,
            known_sheets=known_sheets,
        )
        if observed_address is None:
            kind = (
                FormulaChangeKind.REPLACED
                if literal_address is not None
                else FormulaChangeKind.REMOVED
            )
            changes.append(
                FormulaChange(
                    kind=kind,
                    expected_address=expected_address,
                    observed_address=literal_address,
                    expected_formula=expected_formula.raw,
                    expected_normalized=expected_parsed.normalized,
                    message=_change_message(kind, expected_address, literal_address),
                )
            )
            continue

        matched_observed.add(observed_address)
        observed_formula = observed.formulas[observed_address]
        observed_parsed = parse_formula(
            observed_formula.raw,
            address=observed_address,
            mappings=mappings,
            mapping_direction=MappingDirection.OBSERVED_TO_EXPECTED,
            known_sheets=known_sheets,
        )
        if expected_parsed.normalized != observed_parsed.normalized:
            changes.append(
                FormulaChange(
                    kind=FormulaChangeKind.MODIFIED,
                    expected_address=expected_address,
                    observed_address=observed_address,
                    expected_formula=expected_formula.raw,
                    observed_formula=observed_formula.raw,
                    expected_normalized=expected_parsed.normalized,
                    observed_normalized=observed_parsed.normalized,
                    message=_change_message(
                        FormulaChangeKind.MODIFIED,
                        expected_address,
                        observed_address,
                    ),
                )
            )

    for observed_address, observed_formula in sorted(observed.formulas.items()):
        if observed_address in matched_observed:
            continue
        observed_parsed = parse_formula(
            observed_formula.raw,
            address=observed_address,
            mappings=mappings,
            mapping_direction=MappingDirection.OBSERVED_TO_EXPECTED,
            known_sheets=known_sheets,
        )
        changes.append(
            FormulaChange(
                kind=FormulaChangeKind.ADDED,
                observed_address=observed_address,
                observed_formula=observed_formula.raw,
                observed_normalized=observed_parsed.normalized,
                message=_change_message(
                    FormulaChangeKind.ADDED, None, observed_address
                ),
            )
        )

    for observed_address, observed_formula in sorted(observed.formulas.items()):
        parsed = parse_formula(
            observed_formula.raw,
            address=observed_address,
            mappings=mappings,
            mapping_direction=MappingDirection.OBSERVED_TO_EXPECTED,
            known_sheets=known_sheets,
        )
        mapping = _mapping_for(observed_address.sheet, mappings)
        if mapping is None:
            canonical_expected = observed_address
        else:
            expected_row = mapping.observed_row(observed_address.row)
            expected_column = mapping.observed_column(observed_address.column)
            canonical_expected = (
                CellAddress(
                    observed_address.sheet,
                    expected_row,
                    expected_column,
                )
                if expected_row is not None and expected_column is not None
                else None
            )
        expected_broken = bool(
            canonical_expected is not None
            and expected.formulas.get(canonical_expected)
            and expected.formulas[canonical_expected].has_ref_error
        )
        if parsed.has_ref_error and not expected_broken:
            changes.append(
                FormulaChange(
                    kind=FormulaChangeKind.BROKEN_REFERENCE,
                    observed_address=observed_address,
                    observed_formula=observed_formula.raw,
                    observed_normalized=parsed.normalized,
                    message=_change_message(
                        FormulaChangeKind.BROKEN_REFERENCE, None, observed_address
                    ),
                )
            )

    kind_order = {kind: position for position, kind in enumerate(FormulaChangeKind)}
    changes.sort(
        key=lambda change: (
            (change.address.sheet.casefold() if change.address else ""),
            (change.address.row if change.address else 0),
            (change.address.column if change.address else 0),
            kind_order[change.kind],
        )
    )
    return tuple(changes)


def compare_formulas(
    expected_source: Any,
    observed_source: Any,
    *,
    mappings: Mapping[str, SheetCoordinateMapping] | None = None,
) -> tuple[FormulaChange, ...]:
    """Extract and compare formulas from workbook objects or file paths."""

    expected = extract_formulas(expected_source)
    observed = extract_formulas(observed_source)
    return compare_formula_inventories(expected, observed, mappings=mappings)


def build_inter_sheet_dependencies(
    inventory: FormulaInventory,
) -> tuple[SheetDependency, ...]:
    """Build aggregated static inter-sheet dependencies.

    External workbook references and 3-D references are not expanded.  Missing
    internal target sheets are retained with ``target_exists=False``.
    """

    actual_sheets = {name.casefold(): name for name in inventory.sheet_names}
    cells: dict[tuple[str, str], set[CellAddress]] = defaultdict(set)
    counts: dict[tuple[str, str], int] = defaultdict(int)
    existence: dict[tuple[str, str], bool] = {}

    for address, formula in inventory.formulas.items():
        for reference in formula.references:
            if (
                not reference.explicit_sheet
                or not reference.sheet
                or reference.is_external
                or reference.kind == "three_dimensional"
            ):
                continue
            target = actual_sheets.get(reference.sheet.casefold(), reference.sheet)
            if target.casefold() == address.sheet.casefold():
                continue
            key = (address.sheet, target)
            cells[key].add(address)
            counts[key] += 1
            existence[key] = target.casefold() in actual_sheets

    return tuple(
        SheetDependency(
            source_sheet=source,
            target_sheet=target,
            formula_cells=tuple(sorted(cells[(source, target)])),
            reference_count=counts[(source, target)],
            target_exists=existence[(source, target)],
        )
        for source, target in sorted(
            cells, key=lambda item: (item[0].casefold(), item[1].casefold())
        )
    )


def build_dependency_graph(
    inventory: FormulaInventory,
    *,
    include_missing: bool = True,
) -> dict[str, frozenset[str]]:
    """Return ``source_sheet -> target_sheets`` for static dependencies."""

    graph: dict[str, set[str]] = {
        sheet: set() for sheet in inventory.sheet_names
    }
    for dependency in build_inter_sheet_dependencies(inventory):
        if include_missing or dependency.target_exists:
            graph.setdefault(dependency.source_sheet, set()).add(
                dependency.target_sheet
            )
    return {sheet: frozenset(targets) for sheet, targets in graph.items()}


__all__ = [
    "CellAddress",
    "FormulaChange",
    "FormulaChangeKind",
    "FormulaInventory",
    "FormulaReference",
    "MappingDirection",
    "ParsedFormula",
    "SheetCoordinateMapping",
    "SheetDependency",
    "build_dependency_graph",
    "build_inter_sheet_dependencies",
    "compare_formula_inventories",
    "compare_formulas",
    "extract_formulas",
    "normalize_formula_references",
    "parse_formula",
]
