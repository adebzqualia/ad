from __future__ import annotations

import math
from dataclasses import dataclass
from typing import Any, Iterator, Mapping

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

from .config import AnalysisConfig, SheetRule
from .workbook import CellFeature, SheetSnapshot


@dataclass(frozen=True, slots=True)
class ValueDifference:
    code: str
    expected_coordinate: tuple[int, int]
    observed_coordinate: tuple[int, int]
    expected: Any
    observed: Any
    severity: str

    @property
    def expected_address(self) -> str:
        row, column = self.expected_coordinate
        return f"{get_column_letter(column)}{row}"

    @property
    def observed_address(self) -> str:
        row, column = self.observed_coordinate
        return f"{get_column_letter(column)}{row}"


def _is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def values_equivalent(
    expected: Any,
    observed: Any,
    *,
    absolute_tolerance: float = 0.0,
    relative_tolerance: float = 0.0,
) -> bool:
    """Compare typed values without conflating blanks, booleans and identifiers."""

    if _is_number(expected) and _is_number(observed):
        try:
            return math.isclose(
                float(expected),
                float(observed),
                abs_tol=absolute_tolerance,
                rel_tol=relative_tolerance,
            )
        except (TypeError, ValueError, OverflowError):
            return expected == observed
    if type(expected) is not type(observed):
        return False
    return expected == observed


def _controlled_coordinates(
    rule: SheetRule,
    analysis: AnalysisConfig,
) -> Iterator[tuple[int, int]]:
    seen: set[tuple[int, int]] = set()
    for configured_range in rule.controlled_ranges + rule.critical_ranges:
        min_column, min_row, max_column, max_row = range_boundaries(configured_range)
        max_row = min(max_row, analysis.max_rows)
        max_column = min(max_column, analysis.max_columns)
        for row in range(min_row, max_row + 1):
            for column in range(min_column, max_column + 1):
                coordinate = (row, column)
                if coordinate in seen:
                    continue
                if len(seen) >= analysis.max_cells_per_sheet:
                    raise ValueError(
                        "Les plages contrôlées dépassent "
                        f"analysis.max_cells_per_sheet ({analysis.max_cells_per_sheet})."
                    )
                seen.add(coordinate)
                yield coordinate


def _literal_value(feature: CellFeature | None) -> Any:
    if feature is None:
        return None
    return feature.value


def compare_controlled_values(
    expected: SheetSnapshot,
    observed: SheetSnapshot,
    row_mapping: Mapping[int, int],
    column_mapping: Mapping[int, int],
    rule: SheetRule,
    analysis: AnalysisConfig,
) -> list[ValueDifference]:
    """Compare configured controlled cells in canonical expected coordinates.

    Cells whose row or column disappeared are intentionally skipped: their
    absence is already explained by the structural root cause. Formula cells
    are delegated to the formula comparator so cached results never become
    value anomalies.
    """

    if not analysis.compare_controlled_values:
        return []
    differences: list[ValueDifference] = []
    for expected_row, expected_column in sorted(_controlled_coordinates(rule, analysis)):
        if rule.cell_is_editable(expected_row, expected_column):
            continue
        if not rule.cell_is_monitored(expected_row, expected_column):
            continue
        observed_row = row_mapping.get(expected_row)
        observed_column = column_mapping.get(expected_column)
        if observed_row is None or observed_column is None:
            continue
        expected_feature = expected.cells.get((expected_row, expected_column))
        observed_feature = observed.cells.get((observed_row, observed_column))
        if (expected_feature and expected_feature.raw_formula) or (
            observed_feature and observed_feature.raw_formula
        ):
            continue
        expected_value = _literal_value(expected_feature)
        observed_value = _literal_value(observed_feature)
        if values_equivalent(
            expected_value,
            observed_value,
            absolute_tolerance=analysis.numeric_absolute_tolerance,
            relative_tolerance=analysis.numeric_relative_tolerance,
        ):
            continue
        if expected_value is None:
            code = "VALUE_ADDED_OUTSIDE_EDITABLE_ZONE"
        elif observed_value is None:
            code = "STRUCTURAL_VALUE_REMOVED"
        else:
            code = "CONTROLLED_VALUE_CHANGED"
        differences.append(
            ValueDifference(
                code=code,
                expected_coordinate=(expected_row, expected_column),
                observed_coordinate=(observed_row, observed_column),
                expected=expected_value,
                observed=observed_value,
                severity=(
                    "error"
                    if rule.cell_is_critical(expected_row, expected_column)
                    else "warning"
                ),
            )
        )
    return differences


def display_value(value: Any, max_length: int = 240) -> str:
    if value is None:
        return "Absent"
    rendered = str(value)
    return rendered if len(rendered) <= max_length else rendered[: max_length - 1] + "…"
