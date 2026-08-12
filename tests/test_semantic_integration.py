from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from shutil import copyfile

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.workbook.defined_name import DefinedName

from popscheck import compare_workbooks
from popscheck.config import AnalysisConfig, AppConfig, SheetRule
from popscheck.models import Anomaly, CountryResult
from popscheck.reporting import render_country_html
from tests.helpers import BASE_COLUMNS, BASE_ROWS, create_workbook


DATA_SHEET = "Données"


class SemanticComparisonIntegrationTests(unittest.TestCase):
    def setUp(self) -> None:
        temporary = tempfile.TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        self.root = Path(temporary.name)

    def _pair(
        self,
        expected_value,
        observed_value,
        *,
        coordinate=(DATA_SHEET, "Revenus", "Janvier"),
        config: AppConfig | None = None,
    ):
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET,),
            entries={coordinate: expected_value},
        )
        received = create_workbook(
            self.root / "received" / "France.xlsx",
            sheets=(DATA_SHEET,),
            entries={coordinate: observed_value},
        )
        return compare_workbooks(reference, received, config=config)

    def test_formula_change_is_reported_after_structural_alignment(self) -> None:
        result = self._pair("=C2+D2", "=C2-D2")

        self.assertEqual(
            ["FORMULA_LOGIC_CHANGED"],
            [item.code for item in result.anomalies],
            result.to_dict(),
        )
        anomaly = result.anomalies[0]
        self.assertEqual("formules", anomaly.category)
        self.assertEqual("'Données'!B2", anomaly.location)
        self.assertEqual("B2", anomaly.expected_position)
        self.assertEqual("B2", anomaly.observed_position)
        self.assertTrue(anomaly.id)
        self.assertTrue(anomaly.action.get("country_message"))

    def test_removed_and_replaced_formula_are_distinguished(self) -> None:
        cases = (
            (None, "FORMULA_REMOVED", "Absent"),
            (125, "FORMULA_REPLACED_BY_VALUE", "125"),
        )
        for observed, code, display in cases:
            with self.subTest(code=code):
                result = self._pair("=C2+D2", observed)
                self.assertEqual([code], [item.code for item in result.anomalies])
                self.assertEqual(display, result.anomalies[0].observed)
                self.assertEqual("error", result.anomalies[0].severity)
                self.assertEqual("B2", result.anomalies[0].observed_position)

    def test_added_formula_is_detected_in_a_surviving_cell(self) -> None:
        result = self._pair(None, "=C2+D2")

        self.assertEqual(["FORMULA_ADDED"], [item.code for item in result.anomalies])
        self.assertEqual("Aucune formule", result.anomalies[0].expected)

    def test_shifted_equivalent_formula_does_not_create_formula_noise(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET,),
            entries={(DATA_SHEET, "Marge", "Avril"): "=B4+$A$1"},
        )
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            worksheet = workbook[DATA_SHEET]
            worksheet.delete_rows(3, 1)
            worksheet.delete_cols(3, 1)
            # Expected E4 maps to received D3; the relative reference follows
            # the same semantic row after both structural shifts.
            worksheet["D3"] = "=B3+$A$1"
            workbook.save(received)
        finally:
            workbook.close()

        result = compare_workbooks(reference, received)

        self.assertCountEqual(
            ["ROW_REMOVED", "COLUMN_REMOVED"],
            [item.code for item in result.anomalies],
            result.to_dict(),
        )
        self.assertFalse(
            any(item.code.startswith("FORMULA_") for item in result.anomalies)
        )

    def test_shifted_reference_beyond_structural_extent_is_equivalent(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET,),
        )
        workbook = load_workbook(reference)
        try:
            workbook[DATA_SHEET]["B2"] = "=A100"
            workbook.save(reference)
        finally:
            workbook.close()

        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            worksheet = workbook[DATA_SHEET]
            worksheet.insert_rows(3, 1)
            # openpyxl does not rewrite formulas; this is what Excel saves
            # after inserting a row above the distant referenced cell.
            worksheet["B2"] = "=A101"
            workbook.save(received)
        finally:
            workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(["ROW_ADDED"], [item.code for item in result.anomalies])

    def test_broken_reference_is_a_consequence_of_deleted_column(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=("Forecast", "Summary"),
            entries={("Summary", "Revenus", "Janvier"): "='Forecast'!D2"},
        )
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            workbook["Forecast"].delete_cols(4, 1)
            workbook["Summary"]["B2"] = "=#REF!"
            workbook.save(received)
        finally:
            workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(
            ["COLUMN_REMOVED"],
            [item.code for item in result.anomalies],
            result.to_dict(),
        )
        root = result.anomalies[0]
        self.assertEqual("Forecast", root.sheet)
        self.assertEqual("D", root.expected_position)
        self.assertEqual("error", root.severity)
        by_code = {item["code"]: item for item in root.consequences}
        self.assertIn("INVALID_REFERENCE", by_code)
        self.assertEqual("confirmed", by_code["INVALID_REFERENCE"]["certainty"])
        self.assertIn("'Summary'!B2", by_code["INVALID_REFERENCE"]["sample_locations"])
        self.assertIn("DEPENDENCY_REMOVED", by_code)

        rendered = render_country_html(result)
        self.assertIn("Conséquences rattachées à cette cause", rendered)
        self.assertIn("Action opérationnelle recommandée", rendered)
        self.assertIn("Message pays", rendered)

    def test_missing_referenced_sheet_stays_under_formula_root(self) -> None:
        result = self._pair("='Données'!C2", "='Missing sheet'!C2")

        self.assertEqual(
            ["FORMULA_LOGIC_CHANGED"],
            [item.code for item in result.anomalies],
            result.to_dict(),
        )
        codes = {item["code"] for item in result.anomalies[0].consequences}
        self.assertIn("DEPENDENCY_CHANGED", codes)
        self.assertIn("MISSING_REFERENCED_OBJECT", codes)
        self.assertEqual("error", result.anomalies[0].severity)

    def test_controlled_value_changes_but_default_input_does_not(self) -> None:
        controlled = AppConfig(
            sheet_rules=(
                SheetRule(pattern=DATA_SHEET, controlled_ranges=("B2:B2",)),
            )
        )
        controlled_result = self._pair(100, 101, config=controlled)
        default_result = self._pair(100, 101)

        self.assertEqual(
            ["CONTROLLED_VALUE_CHANGED"],
            [item.code for item in controlled_result.anomalies],
            controlled_result.to_dict(),
        )
        self.assertEqual([], default_result.anomalies, default_result.to_dict())

    def test_external_link_change_is_nested_under_formula_change(self) -> None:
        result = self._pair(
            "='[old.xlsx]Sheet1'!A1",
            "='[new.xlsx]Sheet1'!A1",
        )

        self.assertEqual(
            ["FORMULA_LOGIC_CHANGED"],
            [item.code for item in result.anomalies],
            result.to_dict(),
        )
        self.assertIn(
            "EXTERNAL_LINK_CHANGED",
            {item["code"] for item in result.anomalies[0].consequences},
        )

    def test_unchanged_edge_to_deleted_sheet_is_attached_to_sheet_root(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=("Source", "Inputs"),
            entries={("Source", "Revenus", "Janvier"): "='Inputs'!B2"},
        )
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            del workbook["Inputs"]
            workbook.save(received)
        finally:
            workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(["SHEET_REMOVED"], [item.code for item in result.anomalies])
        root = result.anomalies[0]
        by_code = {item["code"]: item for item in root.consequences}
        self.assertIn("MISSING_REFERENCED_OBJECT", by_code)
        self.assertNotIn("DEPENDENCY_POTENTIALLY_AFFECTED", by_code)

    def test_formula_and_dependency_switches_are_independent(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs", "Other"),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "='Inputs'!B2"},
        )
        received = create_workbook(
            self.root / "received" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs", "Other"),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "='Other'!B2"},
        )

        dependencies_only = compare_workbooks(
            reference,
            received,
            config=AppConfig(
                analysis=AnalysisConfig(
                    compare_formulas=False,
                    compare_dependencies=True,
                )
            ),
        )
        self.assertEqual(
            ["DEPENDENCY_CHANGED"],
            [item.code for item in dependencies_only.anomalies],
        )

        formulas_only = compare_workbooks(
            reference,
            received,
            config=AppConfig(
                analysis=AnalysisConfig(
                    compare_formulas=True,
                    compare_dependencies=False,
                )
            ),
        )
        self.assertEqual(
            ["FORMULA_LOGIC_CHANGED"],
            [item.code for item in formulas_only.anomalies],
        )
        self.assertFalse(formulas_only.anomalies[0].consequences)

    def test_dependencies_in_formula_allowed_zone_are_ignored(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs", "Other"),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "='Inputs'!B2"},
        )
        received = create_workbook(
            self.root / "received" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs", "Other"),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "='Other'!B2"},
        )
        config = AppConfig(
            sheet_rules=(
                SheetRule(
                    pattern=DATA_SHEET,
                    formula_allowed_ranges=("B2:B2",),
                ),
            )
        )

        result = compare_workbooks(reference, received, config=config)

        self.assertFalse(result.anomalies, result.to_dict())

    def test_editable_values_do_not_authorize_external_formulas(self) -> None:
        config = AppConfig(
            sheet_rules=(
                SheetRule(pattern=DATA_SHEET, editable_ranges=("B2:B2",)),
            )
        )

        result = self._pair(100, "='[evil.xlsx]S'!A1", config=config)

        self.assertEqual(["FORMULA_ADDED"], [item.code for item in result.anomalies])
        self.assertIn(
            "EXTERNAL_LINK_ADDED",
            {item["code"] for item in result.anomalies[0].consequences},
        )

    def test_added_formula_uses_canonical_coordinate_for_criticality(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx", sheets=(DATA_SHEET,)
        )
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            worksheet = workbook[DATA_SHEET]
            worksheet.insert_rows(2, 1)
            worksheet["B4"] = "=1+1"
            workbook.save(received)
        finally:
            workbook.close()
        config = AppConfig(
            sheet_rules=(
                SheetRule(pattern=DATA_SHEET, critical_ranges=("B3:B3",)),
            )
        )

        result = compare_workbooks(reference, received, config=config)

        formula = next(item for item in result.anomalies if item.code == "FORMULA_ADDED")
        self.assertEqual("B3", formula.expected_position)
        self.assertEqual("B4", formula.observed_position)
        self.assertEqual("error", formula.severity)

    def test_formula_range_crossing_inserted_row_is_probable_root_impact(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=("Forecast", "Summary"),
            entries={("Summary", "Revenus", "Janvier"): "=SUM('Forecast'!B2:B5)"},
        )
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            workbook["Forecast"].insert_rows(3, 1)
            workbook["Summary"]["B2"] = "=SUM('Forecast'!B2:B6)"
            workbook.save(received)
        finally:
            workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(["ROW_ADDED"], [item.code for item in result.anomalies])
        impact = next(
            item
            for item in result.anomalies[0].consequences
            if item["code"] == "DEPENDENCY_POTENTIALLY_AFFECTED"
        )
        self.assertEqual("probable", impact["certainty"])
        self.assertIn("'Summary'!B2", impact["sample_locations"])

    def test_impact_summary_uses_full_unique_count_not_samples(self) -> None:
        result = CountryResult(key="fr", country="France")
        result.anomalies.append(
            Anomaly(
                category="colonnes",
                code="COLUMN_REMOVED",
                message="Colonne supprimée",
                consequences=[
                    {
                        "id": "impact-1",
                        "certainty": "probable",
                        "count": 12,
                        "unique_target_count": 12,
                        "sample_locations": ["'Summary'!B2"],
                    }
                ],
            )
        )

        self.assertEqual(12, result.to_dict()["impact_summary"]["probable"])

    def test_removed_defined_name_is_a_missing_referenced_object(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET,),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "=Rate"},
        )
        workbook = load_workbook(reference)
        try:
            workbook.defined_names.add(
                DefinedName("Rate", attr_text=f"'{DATA_SHEET}'!$C$2")
            )
            workbook.save(reference)
        finally:
            workbook.close()
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            del workbook.defined_names["Rate"]
            workbook.save(received)
        finally:
            workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(
            ["MISSING_REFERENCED_OBJECT"],
            [item.code for item in result.anomalies],
            result.to_dict(),
        )
        self.assertIn("nom défini", str(result.anomalies[0].expected).casefold())

    def test_changed_defined_name_target_is_a_dependency_change(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET,),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "=Rate"},
        )
        received = create_workbook(
            self.root / "received" / "France.xlsx",
            sheets=(DATA_SHEET,),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "=Rate"},
        )
        for path, target in ((reference, "$C$2"), (received, "$D$2")):
            workbook = load_workbook(path)
            try:
                workbook.defined_names.add(
                    DefinedName(
                        "Rate",
                        attr_text=f"'{DATA_SHEET}'!{target}",
                    )
                )
                workbook.save(path)
            finally:
                workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(["DEPENDENCY_CHANGED"], [item.code for item in result.anomalies])

    def test_defined_name_target_introducing_external_link_is_explicit(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs"),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "=Rate"},
        )
        received = create_workbook(
            self.root / "received" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs"),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "=Rate"},
        )
        for path, target in (
            (reference, "'Inputs'!$B$2"),
            (received, "'[evil.xlsx]S'!$A$1"),
        ):
            workbook = load_workbook(path)
            try:
                workbook.defined_names.add(
                    DefinedName("Rate", attr_text=target)
                )
                workbook.save(path)
            finally:
                workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(["DEPENDENCY_CHANGED"], [item.code for item in result.anomalies])
        external_link = next(
            item
            for item in result.anomalies[0].consequences
            if item["code"] == "EXTERNAL_LINK_ADDED"
        )
        self.assertEqual("confirmed", external_link["certainty"])
        self.assertEqual(1, external_link["count"])
        self.assertIn("evil.xlsx", external_link["explanation"])

    def test_removed_table_used_by_formula_is_not_an_axis_deletion(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx", sheets=(DATA_SHEET,)
        )
        workbook = load_workbook(reference)
        try:
            worksheet = workbook[DATA_SHEET]
            worksheet["A1"] = "Label"
            worksheet["B1"] = "Amount"
            worksheet["D2"] = "=SUM(Table1[Amount])"
            worksheet.add_table(Table(displayName="Table1", ref="A1:B3"))
            workbook.save(reference)
        finally:
            workbook.close()
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            del workbook[DATA_SHEET].tables["Table1"]
            workbook.save(received)
        finally:
            workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(
            ["MISSING_REFERENCED_OBJECT"],
            [item.code for item in result.anomalies],
            result.to_dict(),
        )

    def test_changed_table_range_is_a_dependency_change(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx", sheets=(DATA_SHEET,)
        )
        received = create_workbook(
            self.root / "received" / "France.xlsx", sheets=(DATA_SHEET,)
        )
        for path, table_range in ((reference, "A1:B3"), (received, "A1:B4")):
            workbook = load_workbook(path)
            try:
                worksheet = workbook[DATA_SHEET]
                worksheet["A1"] = "Label"
                worksheet["B1"] = "Amount"
                worksheet["D2"] = "=SUM(Table1[Amount])"
                worksheet.add_table(Table(displayName="Table1", ref=table_range))
                workbook.save(path)
            finally:
                workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(["DEPENDENCY_CHANGED"], [item.code for item in result.anomalies])

    def test_dependency_only_mode_detects_partial_and_coordinate_changes(self) -> None:
        dependency_only = AppConfig(
            analysis=AnalysisConfig(
                compare_formulas=False,
                compare_dependencies=True,
            )
        )
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs"),
            entries={
                (DATA_SHEET, "Revenus", "Janvier"): "='Inputs'!A1",
                (DATA_SHEET, "Coûts", "Janvier"): "='Inputs'!A1",
            },
        )
        received = create_workbook(
            self.root / "received" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs"),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "='Inputs'!A1"},
        )

        removed = compare_workbooks(reference, received, config=dependency_only)

        self.assertEqual(["DEPENDENCY_REMOVED"], [item.code for item in removed.anomalies])

        redirected = create_workbook(
            self.root / "received" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs"),
            entries={
                (DATA_SHEET, "Revenus", "Janvier"): "='Inputs'!B1",
                (DATA_SHEET, "Coûts", "Janvier"): "='Inputs'!A1",
            },
        )
        changed = compare_workbooks(reference, redirected, config=dependency_only)
        self.assertEqual(["DEPENDENCY_CHANGED"], [item.code for item in changed.anomalies])

    def test_dynamic_dependency_is_possible_only_for_relevant_sheet(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs", "Other"),
            entries={
                (DATA_SHEET, "Revenus", "Janvier"): "=INDIRECT(\"'Inputs'!B2\")"
            },
        )

        def compare_deleted_column(sheet_name: str):
            received = self.root / "received" / f"{sheet_name}.xlsx"
            received.parent.mkdir(parents=True, exist_ok=True)
            copyfile(reference, received)
            workbook = load_workbook(received)
            try:
                workbook[sheet_name].delete_cols(2, 1)
                workbook.save(received)
            finally:
                workbook.close()
            return compare_workbooks(reference, received)

        related = compare_deleted_column("Inputs")
        unrelated = compare_deleted_column("Other")

        self.assertIn(
            "DYNAMIC_OR_OPAQUE_DEPENDENCY",
            {item["code"] for item in related.anomalies[0].consequences},
        )
        self.assertFalse(unrelated.anomalies[0].consequences)

    def test_invalid_and_external_formulas_in_added_row_stay_under_row_root(self) -> None:
        cases = (
            ("=#REF!", "INVALID_REFERENCE"),
            ("='[evil.xlsx]S'!A1", "EXTERNAL_LINK_ADDED"),
        )
        for formula, consequence_code in cases:
            with self.subTest(consequence=consequence_code):
                reference = create_workbook(
                    self.root / "sent" / f"{consequence_code}.xlsx",
                    sheets=(DATA_SHEET,),
                )
                received = self.root / "received" / f"{consequence_code}.xlsx"
                received.parent.mkdir(parents=True, exist_ok=True)
                copyfile(reference, received)
                workbook = load_workbook(received)
                try:
                    worksheet = workbook[DATA_SHEET]
                    worksheet.insert_rows(4, 1)
                    worksheet["B4"] = formula
                    workbook.save(received)
                finally:
                    workbook.close()

                result = compare_workbooks(reference, received)

                self.assertEqual(["ROW_ADDED"], [item.code for item in result.anomalies])
                codes = {item["code"] for item in result.anomalies[0].consequences}
                self.assertIn(consequence_code, codes)

    def test_broken_reference_already_present_in_template_is_not_an_anomaly(self) -> None:
        result = self._pair("=#REF!", "=#REF!")

        self.assertFalse(result.anomalies, result.to_dict())

    def test_multi_root_impacts_count_unique_affected_formulas(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=("Inputs", "Summary"),
            entries={
                ("Summary", "Revenus", "Janvier"): "='Inputs'!B2+'Inputs'!D2"
            },
        )
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            workbook["Inputs"].delete_cols(4, 1)
            workbook["Inputs"].delete_cols(2, 1)
            workbook["Summary"]["B2"] = "=#REF!+#REF!"
            workbook.save(received)
        finally:
            workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(
            2,
            sum(item.code == "COLUMN_REMOVED" for item in result.anomalies),
            result.to_dict(),
        )
        self.assertEqual(
            1,
            result.impact_summary["confirmed"],
            "The same affected formula must not be counted once per root cause.",
        )

    def test_dependency_only_change_uses_canonical_critical_severity(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs", "Other"),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "='Inputs'!B2"},
        )
        received = create_workbook(
            self.root / "received" / "France.xlsx",
            sheets=(DATA_SHEET, "Inputs", "Other"),
            entries={(DATA_SHEET, "Revenus", "Janvier"): "='Other'!B2"},
        )
        config = AppConfig(
            analysis=AnalysisConfig(
                compare_formulas=False,
                compare_dependencies=True,
            ),
            sheet_rules=(
                SheetRule(pattern=DATA_SHEET, critical_ranges=("B2:B2",)),
            ),
        )

        result = compare_workbooks(reference, received, config=config)

        self.assertEqual(["DEPENDENCY_CHANGED"], [item.code for item in result.anomalies])
        self.assertEqual("error", result.anomalies[0].severity)

    def test_ignored_sheet_is_excluded_from_audited_metadata(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(DATA_SHEET, "Ignored"),
            entries={("Ignored", "Revenus", "Janvier"): "=1+1"},
        )
        received = create_workbook(
            self.root / "received" / "France.xlsx",
            sheets=(DATA_SHEET, "Ignored"),
            entries={("Ignored", "Revenus", "Janvier"): "=1+2"},
        )
        config = AppConfig(
            sheet_rules=(SheetRule(pattern="Ignored", ignore=True),)
        )

        result = compare_workbooks(reference, received, config=config)

        self.assertFalse(result.anomalies, result.to_dict())
        self.assertEqual(1, result.metadata["analyzed_sheets"])
        formula_analysis = result.metadata["formula_analysis"]
        self.assertEqual(0, formula_analysis["change_count"])
        self.assertEqual(1, formula_analysis["raw_change_count"])

    def test_removed_sheet_inside_three_dimensional_reference_is_possible_impact(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=("Jan", "Feb", "Mar", "Summary"),
            entries={("Summary", "Revenus", "Janvier"): "=SUM(Jan:Mar!B2)"},
        )
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            del workbook["Feb"]
            workbook.save(received)
        finally:
            workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(["SHEET_REMOVED"], [item.code for item in result.anomalies])
        consequences = result.anomalies[0].consequences
        possible = next(
            item
            for item in consequences
            if item["code"] == "DYNAMIC_OR_OPAQUE_DEPENDENCY"
        )
        self.assertEqual("possible", possible["certainty"])
        self.assertEqual("partial", result.metadata["formula_analysis"]["status"])
        self.assertTrue(
            any("3-D reference" in warning for warning in result.warnings),
            result.to_dict(),
        )


if __name__ == "__main__":
    unittest.main()
