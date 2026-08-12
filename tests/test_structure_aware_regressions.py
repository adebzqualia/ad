from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from shutil import copyfile
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table

from popscheck import compare_workbooks
from popscheck.config import AnalysisConfig, AppConfig, SheetRule
from popscheck.models import Status

from tests.helpers import BASE_COLUMNS, BASE_ROWS, create_workbook, without


SHEET_NAME = "Données"


def _replace(items: Sequence[str], old: str, new: str) -> tuple[str, ...]:
    replaced = list(items)
    replaced[replaced.index(old)] = new
    return tuple(replaced)


def _without_many(items: Sequence[str], *removed: str) -> tuple[str, ...]:
    removed_set = set(removed)
    return tuple(item for item in items if item not in removed_set)


def _create_table_workbook(path: Path, headers: Sequence[str]) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Forecast"
    for column, header in enumerate(headers, start=1):
        worksheet.cell(1, column, header)
    for row in range(2, 7):
        worksheet.cell(row, 1, f"Row {row}")
        for column in range(2, len(headers) + 1):
            worksheet.cell(row, column, row * column)
    last_column = get_column_letter(len(headers))
    worksheet.add_table(
        Table(displayName="ForecastTable", ref=f"A1:{last_column}6")
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()
    return path


class StructureAwareRegressionTests(unittest.TestCase):
    def setUp(self) -> None:
        temporary = tempfile.TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        self.root = Path(temporary.name)

    def _compare(
        self,
        *,
        rows: Sequence[str] = BASE_ROWS,
        columns: Sequence[str] = BASE_COLUMNS,
        config=None,
    ):
        sheets = (SHEET_NAME,)
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=sheets,
            rows=BASE_ROWS,
            columns=BASE_COLUMNS,
        )
        received = create_workbook(
            self.root / "received" / "France.xlsx",
            sheets=sheets,
            rows=rows,
            columns=columns,
        )
        return compare_workbooks(reference, received, config=config)

    def _compare_physical_mutation(self, mutation, *, config=None):
        reference = create_workbook(
            self.root / "sent" / "France.xlsx", sheets=(SHEET_NAME,)
        )
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            mutation(workbook[SHEET_NAME])
            workbook.save(received)
        finally:
            workbook.close()
        return compare_workbooks(reference, received, config=config)

    def test_same_position_row_replacement_is_remove_plus_add(self) -> None:
        result = self._compare(
            rows=_replace(BASE_ROWS, "Marge", "Ajustement pays")
        )

        self.assertEqual(Status.ANOMALIES, result.status, result.to_dict())
        self.assertEqual(
            ["ROW_REMOVED", "ROW_ADDED"],
            [anomaly.code for anomaly in result.anomalies],
            result.to_dict(),
        )
        removed, added = result.anomalies
        self.assertEqual(SHEET_NAME, removed.sheet)
        self.assertEqual(4, removed.expected_position)
        self.assertIsNone(removed.observed_position)
        self.assertIn(SHEET_NAME.casefold(), (removed.location or "").casefold())
        self.assertTrue((removed.location or "").endswith("!4:4"))
        self.assertIn("marge", str(removed.expected or "").casefold())
        self.assertEqual("Absent", removed.observed)
        self.assertEqual(SHEET_NAME, added.sheet)
        self.assertIsNone(added.expected_position)
        self.assertEqual(4, added.observed_position)
        self.assertEqual(removed.location, added.location)
        self.assertEqual("Absent", added.expected)
        self.assertIn("ajustement pays", str(added.observed or "").casefold())
        self.assertEqual(2, result.total_anomalies)

    def test_same_position_column_replacement_is_remove_plus_add(self) -> None:
        result = self._compare(
            columns=_replace(BASE_COLUMNS, "Mars", "Trimestre")
        )

        self.assertEqual(Status.ANOMALIES, result.status, result.to_dict())
        self.assertEqual(
            ["COLUMN_REMOVED", "COLUMN_ADDED"],
            [anomaly.code for anomaly in result.anomalies],
            result.to_dict(),
        )
        removed, added = result.anomalies
        self.assertEqual(SHEET_NAME, removed.sheet)
        self.assertEqual("D", removed.expected_position)
        self.assertIsNone(removed.observed_position)
        self.assertIn(SHEET_NAME.casefold(), (removed.location or "").casefold())
        self.assertTrue((removed.location or "").endswith("!D:D"))
        self.assertIn("mars", str(removed.expected or "").casefold())
        self.assertEqual("Absent", removed.observed)
        self.assertEqual(SHEET_NAME, added.sheet)
        self.assertIsNone(added.expected_position)
        self.assertEqual("D", added.observed_position)
        self.assertEqual(removed.location, added.location)
        self.assertEqual("Absent", added.expected)
        self.assertIn("trimestre", str(added.observed or "").casefold())
        self.assertEqual(2, result.total_anomalies)

    def test_simultaneous_row_and_column_deletion_has_no_cascade(self) -> None:
        result = self._compare(
            rows=without(BASE_ROWS, "Marge"),
            columns=without(BASE_COLUMNS, "Mars"),
        )

        self.assertEqual(Status.ANOMALIES, result.status, result.to_dict())
        self.assertEqual(2, len(result.anomalies), result.to_dict())
        self.assertCountEqual(
            ["ROW_REMOVED", "COLUMN_REMOVED"],
            [anomaly.code for anomaly in result.anomalies],
        )
        by_code = {anomaly.code: anomaly for anomaly in result.anomalies}
        self.assertEqual(4, by_code["ROW_REMOVED"].expected_position)
        self.assertEqual("D", by_code["COLUMN_REMOVED"].expected_position)
        self.assertTrue(
            all(anomaly.sheet == SHEET_NAME for anomaly in result.anomalies)
        )
        self.assertTrue(all(anomaly.impact == 1 for anomaly in result.anomalies))
        self.assertEqual(2, result.total_anomalies)

    def test_contiguous_row_deletions_are_one_impacted_range(self) -> None:
        result = self._compare(
            rows=_without_many(BASE_ROWS, "Coûts", "Marge")
        )

        self.assertEqual(Status.ANOMALIES, result.status, result.to_dict())
        self.assertEqual(1, len(result.anomalies), result.to_dict())
        anomaly = result.anomalies[0]
        self.assertEqual("ROW_REMOVED", anomaly.code)
        self.assertEqual(SHEET_NAME, anomaly.sheet)
        self.assertEqual(2, anomaly.impact)
        self.assertIn(SHEET_NAME.casefold(), (anomaly.location or "").casefold())
        self.assertTrue((anomaly.location or "").endswith("!3:4"))
        self.assertEqual(2, result.total_anomalies)
        self.assertEqual(2, result.counts["lignes"])
        serialized = result.to_dict()
        self.assertEqual(1, serialized["root_cause_count"])
        self.assertEqual({"ROW_REMOVED": 2}, serialized["counts_by_code"])
        self.assertEqual("warning", serialized["validation_level"])

    def test_contiguous_column_deletions_are_one_impacted_range(self) -> None:
        result = self._compare(
            columns=_without_many(BASE_COLUMNS, "Février", "Mars")
        )

        self.assertEqual(Status.ANOMALIES, result.status, result.to_dict())
        self.assertEqual(1, len(result.anomalies), result.to_dict())
        anomaly = result.anomalies[0]
        self.assertEqual("COLUMN_REMOVED", anomaly.code)
        self.assertEqual(SHEET_NAME, anomaly.sheet)
        self.assertEqual(2, anomaly.impact)
        self.assertIn(SHEET_NAME.casefold(), (anomaly.location or "").casefold())
        self.assertTrue((anomaly.location or "").endswith("!C:D"))
        self.assertEqual(2, result.total_anomalies)
        self.assertEqual(2, result.counts["colonnes"])

    def test_physical_excel_column_deletion_has_no_trailing_residue_anomaly(self) -> None:
        result = self._compare_physical_mutation(
            lambda worksheet: worksheet.delete_cols(4, 1)
        )

        self.assertEqual(["COLUMN_REMOVED"], [item.code for item in result.anomalies])
        self.assertEqual("D", result.anomalies[0].expected_position)
        summary = result.metadata["sheet_summaries"][SHEET_NAME]
        self.assertEqual(-1, summary["column_delta"])
        self.assertEqual(0, summary["row_delta"])

    def test_physical_excel_row_deletion_has_no_trailing_residue_anomaly(self) -> None:
        result = self._compare_physical_mutation(
            lambda worksheet: worksheet.delete_rows(4, 1)
        )

        self.assertEqual(["ROW_REMOVED"], [item.code for item in result.anomalies])
        self.assertEqual(4, result.anomalies[0].expected_position)
        summary = result.metadata["sheet_summaries"][SHEET_NAME]
        self.assertEqual(-1, summary["row_delta"])
        self.assertEqual(0, summary["column_delta"])

    def test_table_span_change_does_not_cascade_across_the_sheet(self) -> None:
        reference = _create_table_workbook(
            self.root / "sent" / "France.xlsx",
            ("Label", "Jan", "Feb", "Mar", "Apr", "Total"),
        )
        received = _create_table_workbook(
            self.root / "received" / "France.xlsx",
            ("Label", "Jan", "Feb", "Apr", "Total"),
        )

        result = compare_workbooks(reference, received)

        self.assertEqual(["COLUMN_REMOVED"], [item.code for item in result.anomalies])
        self.assertEqual("D", result.anomalies[0].expected_position)

    def test_monitored_range_tracks_an_insertion_past_its_physical_boundary(self) -> None:
        config = AppConfig(
            sheet_rules=(
                SheetRule(pattern=SHEET_NAME, monitored_ranges=("A1:F6",)),
            )
        )
        result = self._compare_physical_mutation(
            lambda worksheet: worksheet.insert_rows(4, 1), config=config
        )

        self.assertEqual(["ROW_ADDED"], [item.code for item in result.anomalies])
        self.assertEqual(4, result.anomalies[0].observed_position)
        self.assertEqual(
            1, result.metadata["sheet_summaries"][SHEET_NAME]["row_delta"]
        )

    def test_monitored_range_row_insertion_does_not_cascade_to_columns(self) -> None:
        config = AppConfig(
            sheet_rules=(
                SheetRule(pattern=SHEET_NAME, monitored_ranges=("A1:F6",)),
            )
        )
        received_rows = (*BASE_ROWS[:3], "Ajustement pays", *BASE_ROWS[3:])

        result = self._compare(rows=received_rows, config=config)

        self.assertEqual(["ROW_ADDED"], [item.code for item in result.anomalies])
        self.assertEqual(4, result.anomalies[0].observed_position)

    def test_monitored_range_column_insertion_does_not_cascade_to_rows(self) -> None:
        config = AppConfig(
            sheet_rules=(
                SheetRule(pattern=SHEET_NAME, monitored_ranges=("A1:F6",)),
            )
        )
        received_columns = (
            *BASE_COLUMNS[:3],
            "Ajustement pays",
            *BASE_COLUMNS[3:],
        )

        result = self._compare(columns=received_columns, config=config)

        self.assertEqual(["COLUMN_ADDED"], [item.code for item in result.anomalies])
        self.assertEqual("D", result.anomalies[0].observed_position)

    def test_removed_critical_column_is_an_error(self) -> None:
        config = AppConfig(
            sheet_rules=(
                SheetRule(pattern=SHEET_NAME, critical_ranges=("D1:D6",)),
            )
        )

        result = self._compare(columns=without(BASE_COLUMNS, "Mars"), config=config)

        self.assertEqual(["COLUMN_REMOVED"], [item.code for item in result.anomalies])
        self.assertEqual("error", result.anomalies[0].severity)

    def test_controlled_range_extends_scan_without_monitored_range(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx", sheets=(SHEET_NAME,)
        )
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            workbook[SHEET_NAME]["B100"] = 123
            workbook.save(received)
        finally:
            workbook.close()
        config = AppConfig(
            sheet_rules=(
                SheetRule(pattern=SHEET_NAME, controlled_ranges=("B100:B100",)),
            )
        )

        result = compare_workbooks(reference, received, config=config)

        self.assertEqual(
            ["VALUE_ADDED_OUTSIDE_EDITABLE_ZONE"],
            [item.code for item in result.anomalies],
            result.to_dict(),
        )

    def test_formula_on_removed_row_does_not_create_dependency_root(self) -> None:
        reference = create_workbook(
            self.root / "sent" / "France.xlsx",
            sheets=(SHEET_NAME, "Inputs"),
            entries={(SHEET_NAME, "Marge", "Janvier"): "='Inputs'!A1"},
        )
        received = self.root / "received" / "France.xlsx"
        received.parent.mkdir(parents=True, exist_ok=True)
        copyfile(reference, received)
        workbook = load_workbook(received)
        try:
            workbook[SHEET_NAME].delete_rows(4, 1)
            workbook.save(received)
        finally:
            workbook.close()

        result = compare_workbooks(reference, received)

        self.assertEqual(["ROW_REMOVED"], [item.code for item in result.anomalies])

    def test_monitored_extent_is_clamped_to_analysis_limit(self) -> None:
        config = AppConfig(
            analysis=AnalysisConfig(max_rows=10),
            sheet_rules=(
                SheetRule(pattern=SHEET_NAME, monitored_ranges=("A1:A100",)),
            ),
        )

        result = self._compare(config=config)

        self.assertFalse(result.anomalies, result.to_dict())
        summary = result.metadata["sheet_summaries"][SHEET_NAME]
        self.assertEqual(10, summary["expected_rows"])
        self.assertEqual(10, summary["observed_rows"])

    def test_noncontiguous_additions_receive_distinct_stable_ids(self) -> None:
        received_rows = (
            BASE_ROWS[0],
            "Ajustement",
            *BASE_ROWS[1:4],
            "Ajustement",
            *BASE_ROWS[4:],
        )

        result = self._compare(rows=received_rows)

        additions = [item for item in result.anomalies if item.code == "ROW_ADDED"]
        self.assertEqual(2, len(additions), result.to_dict())
        self.assertEqual(2, len({item.id for item in additions}))


if __name__ == "__main__":
    unittest.main()
