from __future__ import annotations

import hashlib
import math
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import column_index_from_string, range_boundaries

from .config import AnalysisConfig, SheetRule


class WorkbookLimitError(RuntimeError):
    """Le classeur dépasse une limite de sûreté configurée."""


_A1_REFERENCE_RE = re.compile(r"(?<![A-Z0-9_.])(?P<ref>\$?[A-Z]{1,3}\$?\d+)(?![A-Z0-9_])")
_SPACE_RE = re.compile(r"\s+")


@dataclass(frozen=True, slots=True)
class CellFeature:
    row: int
    column: int
    style: str = ""
    formula: str = ""
    label: str = ""
    value_kind: str = "blank"

    @property
    def has_structural_evidence(self) -> bool:
        return bool(self.style or self.formula)


@dataclass(frozen=True, slots=True)
class StructuralRange:
    kind: str
    min_row: int
    max_row: int
    min_column: int
    max_column: int
    tag: str = ""

    def axis_token(self, axis: str, index: int) -> str | None:
        if axis == "row":
            start, end = self.min_row, self.max_row
            orthogonal_span = self.max_column - self.min_column + 1
        else:
            start, end = self.min_column, self.max_column
            orthogonal_span = self.max_row - self.min_row + 1
        if not start <= index <= end:
            return None
        if start == end:
            role = "single"
        elif index == start:
            role = "start"
        elif index == end:
            role = "end"
        else:
            role = "inside"
        axis_span = end - start + 1
        return f"{self.kind}|{role}|a{axis_span}|o{orthogonal_span}|{self.tag}"


@dataclass(slots=True)
class SheetSnapshot:
    name: str
    index: int
    state: str
    kind: str = "worksheet"
    cells: dict[tuple[int, int], CellFeature] = field(default_factory=dict)
    ranges: list[StructuralRange] = field(default_factory=list)
    row_dimensions: dict[int, str] = field(default_factory=dict)
    column_dimensions: dict[int, str] = field(default_factory=dict)
    content_row_max: int = 0
    content_column_max: int = 0
    formula_row_max: int = 0
    formula_column_max: int = 0
    styled_rows: set[int] = field(default_factory=set)
    styled_columns: set[int] = field(default_factory=set)
    warnings: list[str] = field(default_factory=list)

    @property
    def labels(self) -> set[str]:
        return {feature.label for feature in self.cells.values() if feature.label}

    def extent(
        self,
        axis: str,
        analysis: AnalysisConfig,
        rule: SheetRule,
        *,
        include_literal_values: bool,
    ) -> int:
        configured = rule.monitored_extent()
        if configured:
            return configured[0] if axis == "row" else configured[1]
        if axis == "row":
            content_max = self.content_row_max
            formula_max = self.formula_row_max
            style_indexes = self.styled_rows
            dimension_indexes = set(self.row_dimensions)
            range_max = max(
                (
                    item.max_row
                    for item in self.ranges
                    if not (
                        item.kind == "D"
                        and (item.max_row >= 1_048_576 or item.max_column >= 16_384)
                    )
                ),
                default=0,
            )
            limit = analysis.max_rows
        else:
            content_max = self.content_column_max
            formula_max = self.formula_column_max
            style_indexes = self.styled_columns
            dimension_indexes = set(self.column_dimensions)
            range_max = max(
                (
                    item.max_column
                    for item in self.ranges
                    if not (
                        item.kind == "D"
                        and (item.max_row >= 1_048_576 or item.max_column >= 16_384)
                    )
                ),
                default=0,
            )
            limit = analysis.max_columns

        anchor = max(range_max, formula_max, content_max if include_literal_values else 0)
        supporting_indexes = sorted(style_indexes | dimension_indexes)
        if anchor == 0 and supporting_indexes:
            # Un classeur uniquement formaté reste analysable, mais une trace de
            # style isolée très loin est bornée par la limite de sûreté.
            anchor = supporting_indexes[0]
        current = anchor
        for index in supporting_indexes:
            if index <= current:
                continue
            if index - current <= analysis.max_style_gap:
                current = index
            else:
                break
        return min(max(current, 0), limit)


@dataclass(slots=True)
class WorkbookSnapshot:
    path: Path
    sheet_order: list[str]
    sheets: dict[str, SheetSnapshot]
    warnings: list[str] = field(default_factory=list)


def normalize_label(value: str) -> str:
    normalized = unicodedata.normalize("NFC", value).strip()
    normalized = _SPACE_RE.sub(" ", normalized)
    if not normalized or len(normalized) > 300:
        return ""
    return normalized.casefold()


def _normalize_reference(reference: str, origin_row: int, origin_column: int) -> str:
    column_match = re.match(r"(\$?)([A-Z]{1,3})(\$?)(\d+)$", reference)
    if not column_match:
        return reference
    absolute_column, column_letters, absolute_row, row_text = column_match.groups()
    column = column_index_from_string(column_letters)
    row = int(row_text)
    row_part = str(row) if absolute_row else f"[{row - origin_row:+d}]"
    column_part = str(column) if absolute_column else f"[{column - origin_column:+d}]"
    return f"R{row_part}C{column_part}"


def _normalize_range_token(token: str, origin_row: int, origin_column: int) -> str:
    # Les préfixes de feuille/classeur, noms définis et références structurées
    # sont conservés. Seules les véritables références A1 sont relativisées.
    def replace(match: re.Match[str]) -> str:
        return _normalize_reference(match.group("ref"), origin_row, origin_column)

    return _A1_REFERENCE_RE.sub(replace, token.upper())


def normalize_formula(formula: Any, row: int, column: int) -> str:
    if hasattr(formula, "text"):
        formula = formula.text
    text = str(formula or "")
    if not text:
        return ""
    if not text.startswith("="):
        text = f"={text}"
    try:
        tokenizer = Tokenizer(text)
        parts: list[str] = []
        for token in tokenizer.items:
            value = token.value
            if token.type == "OPERAND" and token.subtype == "RANGE":
                value = _normalize_range_token(value, row, column)
            elif token.type == "WHITE-SPACE":
                value = " "
            parts.append(value)
        normalized = "".join(parts)
    except Exception:
        normalized = _normalize_range_token(text, row, column)
    return normalized.upper()


def _round_float(value: Any) -> Any:
    if isinstance(value, float):
        if math.isnan(value):
            return "nan"
        return round(value, 6)
    return value


def _color_tuple(color: Any) -> tuple[Any, ...]:
    if color is None:
        return ()
    return (
        getattr(color, "type", None),
        getattr(color, "rgb", None),
        getattr(color, "indexed", None),
        getattr(color, "theme", None),
        _round_float(getattr(color, "tint", None)),
        getattr(color, "auto", None),
    )


def _side_tuple(side: Any) -> tuple[Any, ...]:
    if side is None:
        return ()
    return (getattr(side, "style", None), _color_tuple(getattr(side, "color", None)))


def canonical_style(cell: Any) -> str:
    if not getattr(cell, "has_style", False):
        return ""
    font = cell.font
    fill = cell.fill
    border = cell.border
    alignment = cell.alignment
    protection = cell.protection
    payload = (
        (
            font.name,
            _round_float(font.sz),
            font.b,
            font.i,
            font.u,
            font.strike,
            _color_tuple(font.color),
            font.vertAlign,
            font.outline,
            font.shadow,
            font.condense,
            font.extend,
            font.family,
            font.charset,
            font.scheme,
        ),
        (
            fill.fill_type,
            _color_tuple(fill.fgColor),
            _color_tuple(fill.bgColor),
        ),
        (
            _side_tuple(border.left),
            _side_tuple(border.right),
            _side_tuple(border.top),
            _side_tuple(border.bottom),
            _side_tuple(border.diagonal),
            _side_tuple(border.vertical),
            _side_tuple(border.horizontal),
            border.diagonalUp,
            border.diagonalDown,
            border.outline,
        ),
        (
            alignment.horizontal,
            alignment.vertical,
            alignment.textRotation,
            alignment.wrapText,
            alignment.shrinkToFit,
            _round_float(alignment.indent),
            _round_float(alignment.relativeIndent),
            alignment.justifyLastLine,
            alignment.readingOrder,
        ),
        cell.number_format,
        (protection.locked, protection.hidden),
        getattr(cell, "quotePrefix", None),
    )
    return hashlib.blake2b(repr(payload).encode("utf-8"), digest_size=10).hexdigest()


def _dimension_token(dimension: Any, axis: str) -> str:
    if axis == "row":
        payload = (
            _round_float(getattr(dimension, "height", None)),
            getattr(dimension, "hidden", False),
            getattr(dimension, "outlineLevel", 0),
            getattr(dimension, "collapsed", False),
            getattr(dimension, "thickTop", False),
            getattr(dimension, "thickBottom", False),
        )
    else:
        payload = (
            _round_float(getattr(dimension, "width", None)),
            getattr(dimension, "hidden", False),
            getattr(dimension, "outlineLevel", 0),
            getattr(dimension, "collapsed", False),
            getattr(dimension, "bestFit", False),
        )
    return hashlib.blake2b(repr(payload).encode("utf-8"), digest_size=8).hexdigest()


def _safe_tag(value: Any, max_length: int = 100) -> str:
    normalized = _SPACE_RE.sub(" ", str(value or "")).strip()
    return normalized[:max_length].casefold()


def _iter_cell_objects(worksheet: Any) -> Iterable[Any]:
    # Une itération sparse évite de matérialiser toutes les cellules jusqu'à
    # max_row/max_column, souvent gonflés par un formatage résiduel.
    for cell in getattr(worksheet, "_cells", {}).values():
        if not isinstance(cell, MergedCell):
            yield cell


def _extract_worksheet(
    worksheet: Any,
    index: int,
    analysis: AnalysisConfig,
) -> SheetSnapshot:
    cells_raw = getattr(worksheet, "_cells", {})
    if len(cells_raw) > analysis.max_cells_per_sheet:
        raise WorkbookLimitError(
            f"La feuille {worksheet.title!r} contient {len(cells_raw):,} cellules matérialisées "
            f"(limite : {analysis.max_cells_per_sheet:,})."
        )
    snapshot = SheetSnapshot(
        name=worksheet.title,
        index=index,
        state=getattr(worksheet, "sheet_state", "visible"),
    )
    style_cache: dict[int, str] = {}
    for cell in _iter_cell_objects(worksheet):
        row, column = cell.row, cell.column
        if row > analysis.max_rows or column > analysis.max_columns:
            snapshot.warnings.append(
                f"Cellule {cell.coordinate} hors limites d'analyse; elle a été ignorée."
            )
            continue
        value = cell.value
        has_value = value is not None
        if has_value:
            snapshot.content_row_max = max(snapshot.content_row_max, row)
            snapshot.content_column_max = max(snapshot.content_column_max, column)
        formula = ""
        label = ""
        value_kind = "blank"
        if cell.data_type == "f":
            formula = normalize_formula(value, row, column)
            value_kind = "formula"
            snapshot.formula_row_max = max(snapshot.formula_row_max, row)
            snapshot.formula_column_max = max(snapshot.formula_column_max, column)
        elif isinstance(value, str):
            label = normalize_label(value)
            value_kind = "text"
        elif value is not None:
            value_kind = cell.data_type or type(value).__name__

        style = ""
        if cell.has_style:
            style_id = int(getattr(cell, "style_id", 0))
            style = style_cache.get(style_id, "")
            if not style:
                style = canonical_style(cell)
                style_cache[style_id] = style
            snapshot.styled_rows.add(row)
            snapshot.styled_columns.add(column)
        if has_value or style:
            snapshot.cells[(row, column)] = CellFeature(
                row=row,
                column=column,
                style=style,
                formula=formula,
                label=label,
                value_kind=value_kind,
            )

    for merged_range in worksheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        snapshot.ranges.append(
            StructuralRange("M", min_row, max_row, min_col, max_col)
        )

    validations = getattr(getattr(worksheet, "data_validations", None), "dataValidation", [])
    for validation in validations:
        tag = ":".join(
            filter(
                None,
                (
                    _safe_tag(getattr(validation, "type", "")),
                    _safe_tag(getattr(validation, "operator", "")),
                ),
            )
        )
        sqref = getattr(validation, "sqref", None)
        for validation_range in getattr(sqref, "ranges", ()):
            min_col, min_row, max_col, max_row = range_boundaries(str(validation_range))
            snapshot.ranges.append(
                StructuralRange("D", min_row, max_row, min_col, max_col, tag)
            )

    for table in getattr(worksheet, "tables", {}).values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        tag = _safe_tag(getattr(table, "displayName", ""))
        snapshot.ranges.append(
            StructuralRange("T", min_row, max_row, min_col, max_col, tag)
        )

    for row_index, dimension in worksheet.row_dimensions.items():
        if row_index <= analysis.max_rows and (
            dimension.height is not None
            or dimension.hidden
            or dimension.outlineLevel
            or dimension.collapsed
            or dimension.has_style
        ):
            snapshot.row_dimensions[row_index] = _dimension_token(dimension, "row")
            snapshot.styled_rows.add(row_index)

    for dimension in worksheet.column_dimensions.values():
        start = max(1, int(getattr(dimension, "min", 0) or 0))
        end = min(analysis.max_columns, int(getattr(dimension, "max", start) or start))
        if not (
            dimension.width is not None
            or dimension.hidden
            or dimension.outlineLevel
            or dimension.collapsed
            or dimension.has_style
        ):
            continue
        token = _dimension_token(dimension, "column")
        for column_index in range(start, end + 1):
            snapshot.column_dimensions[column_index] = token
            snapshot.styled_columns.add(column_index)
    return snapshot


def load_snapshot(path: str | Path, analysis: AnalysisConfig) -> WorkbookSnapshot:
    workbook_path = Path(path)
    keep_vba = workbook_path.suffix.lower() in {".xlsm", ".xltm"}
    workbook = load_workbook(
        workbook_path,
        read_only=False,
        data_only=False,
        keep_vba=keep_vba,
        keep_links=True,
        rich_text=False,
    )
    try:
        sheet_order = list(workbook.sheetnames)
        sheets: dict[str, SheetSnapshot] = {}
        warnings: list[str] = []
        for index, sheet in enumerate(getattr(workbook, "_sheets", []), start=1):
            if hasattr(sheet, "_cells"):
                snapshot = _extract_worksheet(sheet, index, analysis)
            else:
                snapshot = SheetSnapshot(
                    name=sheet.title,
                    index=index,
                    state=getattr(sheet, "sheet_state", "visible"),
                    kind="chartsheet",
                )
            sheets[snapshot.name] = snapshot
            warnings.extend(f"{snapshot.name}: {warning}" for warning in snapshot.warnings)
        return WorkbookSnapshot(
            path=workbook_path,
            sheet_order=sheet_order,
            sheets=sheets,
            warnings=warnings,
        )
    finally:
        workbook.close()
