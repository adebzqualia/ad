from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_ROWS = ("En-tête", "Revenus", "Coûts", "Marge", "Prévision", "Total")
BASE_COLUMNS = ("Libellé", "Janvier", "Février", "Mars", "Avril", "Total")
BASE_SHEETS = ("Synthèse", "Données", "Prévisions")

EntryKey = tuple[str, str, str]


def _stable_number(value: str, modulo: int) -> int:
    digest = hashlib.sha256(value.encode("utf-8")).digest()
    return int.from_bytes(digest[:4], "big") % modulo


def _argb(value: str) -> str:
    return "FF" + hashlib.sha256(value.encode("utf-8")).hexdigest()[:6].upper()


def create_workbook(
    path: str | Path,
    *,
    sheets: Sequence[str] = BASE_SHEETS,
    rows: Sequence[str] = BASE_ROWS,
    columns: Sequence[str] = BASE_COLUMNS,
    entries: Mapping[EntryKey, object] | None = None,
    fill_inputs: bool = False,
) -> Path:
    """Crée un petit template dont chaque axe a une signature unique et stable.

    Les cellules intérieures sont les zones de saisie. Les libellés, styles,
    dimensions et formules de total servent de preuves structurelles, sans faire
    dépendre les tests des valeurs saisies par un pays.
    """

    if not sheets:
        raise ValueError("Un classeur de test doit contenir au moins une feuille")

    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    values = entries or {}
    workbook = Workbook()

    for sheet_index, sheet_name in enumerate(sheets):
        worksheet = workbook.active if sheet_index == 0 else workbook.create_sheet()
        worksheet.title = sheet_name
        worksheet.freeze_panes = "B2"

        for physical_row, row_id in enumerate(rows, start=1):
            worksheet.row_dimensions[physical_row].height = 18 + _stable_number(
                f"row:{row_id}", 8
            )
            for physical_column, column_id in enumerate(columns, start=1):
                cell = worksheet.cell(physical_row, physical_column)
                is_header_row = physical_row == 1
                is_label_column = physical_column == 1
                is_total = row_id == "Total" or column_id == "Total"

                if is_header_row and is_label_column:
                    cell.value = f"Template::{sheet_name}"
                elif is_header_row:
                    cell.value = f"COL::{column_id}"
                elif is_label_column:
                    cell.value = f"ROW::{row_id}"
                elif is_total:
                    formula_id = (
                        _stable_number(f"sheet:{sheet_name}", 1000) * 1_000_000
                        + _stable_number(f"row:{row_id}", 1000) * 1000
                        + _stable_number(f"column:{column_id}", 1000)
                    )
                    cell.value = f"={formula_id}"
                else:
                    key = (sheet_name, row_id, column_id)
                    if key in values:
                        cell.value = values[key]
                    elif fill_inputs:
                        seed = _stable_number("|".join(key), 10_000)
                        cell.value = f"Saisie {seed}" if seed % 2 else seed / 10

                cell.fill = PatternFill("solid", fgColor=_argb(f"row:{row_id}"))
                cell.font = Font(
                    name="Calibri",
                    size=10 + _stable_number(f"font-size:{row_id}", 3),
                    bold=is_header_row or is_label_column or is_total,
                    color=_argb(f"column:{column_id}"),
                )
                cell.border = Border(
                    bottom=Side(
                        style=("thin", "medium", "dashed")[
                            _stable_number(f"border:{row_id}", 3)
                        ],
                        color="FF606060",
                    )
                )
                cell.alignment = Alignment(
                    horizontal=("left", "center", "right")[
                        _stable_number(f"align:{column_id}", 3)
                    ]
                )
                cell.number_format = (
                    "0.00" if _stable_number(f"format:{row_id}:{column_id}", 2) else "0"
                )

        for physical_column, column_id in enumerate(columns, start=1):
            letter = get_column_letter(physical_column)
            worksheet.column_dimensions[letter].width = 12 + _stable_number(
                f"column:{column_id}", 9
            )

    workbook.save(output)
    workbook.close()
    return output


def insert_after(items: Sequence[str], existing: str, added: str) -> tuple[str, ...]:
    result = list(items)
    result.insert(result.index(existing) + 1, added)
    return tuple(result)


def without(items: Sequence[str], removed: str) -> tuple[str, ...]:
    return tuple(item for item in items if item != removed)


def move_after(items: Sequence[str], moved: str, destination: str) -> tuple[str, ...]:
    result = [item for item in items if item != moved]
    result.insert(result.index(destination) + 1, moved)
    return tuple(result)


def file_fingerprint(path: str | Path) -> tuple[str, int, int]:
    candidate = Path(path)
    stat = candidate.stat()
    digest = hashlib.sha256(candidate.read_bytes()).hexdigest()
    return digest, stat.st_size, stat.st_mtime_ns
