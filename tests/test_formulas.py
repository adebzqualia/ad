from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from popscheck.formulas import (
    CellAddress,
    FormulaChangeKind,
    MappingDirection,
    SheetCoordinateMapping,
    build_dependency_graph,
    build_inter_sheet_dependencies,
    compare_formulas,
    extract_formulas,
    normalize_formula_references,
    parse_formula,
)


class FormulaModuleTests(unittest.TestCase):
    def setUp(self) -> None:
        temporary = tempfile.TemporaryDirectory()
        self.addCleanup(temporary.cleanup)
        self.root = Path(temporary.name)

    def test_parse_formula_ignores_ref_text_but_detects_ref_error(self) -> None:
        text_only = parse_formula('="#REF!"')
        broken = parse_formula("=#REF!")

        self.assertFalse(text_only.has_ref_error)
        self.assertTrue(broken.has_ref_error)

    def test_formatting_spaces_do_not_change_formula_semantics(self) -> None:
        compact = parse_formula("=B1+C1")
        spaced = parse_formula("=B1 + C1")

        self.assertEqual(compact.normalized, spaced.normalized)

    def test_external_reference_keeps_windows_path_identity(self) -> None:
        parsed = parse_formula(r"='C:\folder\[Book.xlsx]Data'!A1")

        self.assertEqual(1, len(parsed.references))
        self.assertEqual(r"C:\folder\Book.xlsx", parsed.references[0].external_book)

    def test_reference_normalization_uses_observed_to_expected_mappings(self) -> None:
        mappings = {
            "Data": SheetCoordinateMapping(
                rows={1: 1, 2: 3, 3: 4},
                columns={1: 1, 2: 2},
            )
        }

        normalized = normalize_formula_references(
            "=A3+$B$4",
            address=CellAddress("Data", 3, 2),
            mappings=mappings,
            mapping_direction=MappingDirection.OBSERVED_TO_EXPECTED,
        )

        self.assertEqual("=A2+$B$3", normalized)

    def test_shifted_formula_is_unchanged_after_mapping(self) -> None:
        expected = Workbook()
        expected.active.title = "Data"
        expected["Data"]["B2"] = "=A2+1"
        observed = Workbook()
        observed.active.title = "Data"
        observed["Data"]["B3"] = "=A3+1"
        self.addCleanup(expected.close)
        self.addCleanup(observed.close)

        changes = compare_formulas(
            expected,
            observed,
            mappings={"Data": SheetCoordinateMapping(rows={1: 1, 2: 3})},
        )

        self.assertEqual((), changes)

    def test_formula_change_categories_and_broken_reference(self) -> None:
        expected = Workbook()
        expected.active.title = "Data"
        expected["Data"]["A1"] = "=B1+1"
        expected["Data"]["A2"] = "=B2"
        expected["Data"]["A4"] = "=B4"
        expected["Data"]["A5"] = "=B5"
        expected["Data"]["A6"] = '="#REF!"'

        observed = Workbook()
        observed.active.title = "Data"
        observed["Data"]["A2"] = 42
        observed["Data"]["A3"] = "=B3"
        observed["Data"]["A4"] = "=B4+1"
        observed["Data"]["A5"] = "=#REF!"
        observed["Data"]["A6"] = '="#REF!"'
        self.addCleanup(expected.close)
        self.addCleanup(observed.close)

        changes = compare_formulas(expected, observed)

        self.assertEqual(
            [
                FormulaChangeKind.REMOVED,
                FormulaChangeKind.REPLACED,
                FormulaChangeKind.ADDED,
                FormulaChangeKind.MODIFIED,
                FormulaChangeKind.MODIFIED,
                FormulaChangeKind.BROKEN_REFERENCE,
            ],
            [change.kind for change in changes],
        )

    def test_static_dependencies_exclude_external_and_retain_missing_sheet(self) -> None:
        workbook = Workbook()
        inputs = workbook.active
        inputs.title = "Inputs"
        summary = workbook.create_sheet("Summary")
        summary["A1"] = (
            "='Inputs'!A1+'Inputs'!A2+'Missing sheet'!B2"
            "+'[Other.xlsx]Inputs'!C3"
        )
        self.addCleanup(workbook.close)

        inventory = extract_formulas(workbook)
        dependencies = build_inter_sheet_dependencies(inventory)
        graph = build_dependency_graph(inventory)

        self.assertEqual(2, len(dependencies))
        by_target = {dependency.target_sheet: dependency for dependency in dependencies}
        self.assertEqual(2, by_target["Inputs"].reference_count)
        self.assertTrue(by_target["Inputs"].target_exists)
        self.assertEqual(1, by_target["Missing sheet"].reference_count)
        self.assertFalse(by_target["Missing sheet"].target_exists)
        self.assertEqual(
            frozenset({"Inputs", "Missing sheet"}), graph["Summary"]
        )

    def test_extract_formulas_accepts_a_path_and_closes_its_workbook(self) -> None:
        path = self.root / "book.xlsx"
        workbook = Workbook()
        workbook.active["A1"] = "=1+1"
        workbook.save(path)
        workbook.close()

        inventory = extract_formulas(path)

        self.assertEqual(1, len(inventory.formulas))
        path.unlink()  # succeeds on Windows only if the extractor closed the file

    def test_extract_formulas_inventories_defined_names(self) -> None:
        workbook = Workbook()
        workbook.defined_names.add(
            DefinedName("Rate", attr_text=f"'{workbook.active.title}'!$A$1")
        )
        self.addCleanup(workbook.close)

        inventory = extract_formulas(workbook)

        self.assertIn("Rate", inventory.defined_names)

    def test_literal_text_starting_with_equals_is_not_promoted_to_formula(self) -> None:
        workbook = Workbook()
        cell = workbook.active["A1"]
        cell.value = "=displayed text"
        cell.data_type = "s"
        self.addCleanup(workbook.close)

        inventory = extract_formulas(workbook)

        address = CellAddress(workbook.active.title, 1, 1)
        self.assertNotIn(address, inventory.formulas)
        self.assertIn(address, inventory.literal_cells)


if __name__ == "__main__":
    unittest.main()
